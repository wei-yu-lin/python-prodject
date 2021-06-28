from YUSCO.Util.mail_sender import SEND_MAIL
import openpyxl
workbook = openpyxl.load_workbook('冷軋運轉日報表.xlsx')
sheet = workbook['Sheet1']
mail = {}
for i in range(2, 902):  # 902
    userno = sheet.cell(row=i, column=3).value
    priv = sheet.cell(row=i, column=2).value
    prod_line = sheet.cell(row=i, column=1).value
    mail.setdefault(userno, []).append('產線:'+prod_line+',權限等級:'+priv)
# print(mail)
for temp in mail:
  msg_body = '<html><head></head>'
  msg_body += '<body>'
  msg_body += ' <b><h>各單位長官好<br>'
  # msg_body += '依總經理指示全面清查程式相關權限<br>'
  # msg_body += '均需提出電子表單(程式權限申請單)申請<br>'
  # msg_body += '請于本週五(2/5)中午前提出，若無提出申請，下周一停止相關程式使用<br>'
  # msg_body += '請配合相關工作<br>'
  # msg_body += '1.目前的人員權限名單，可供整理參考如附件<br>'
  # msg_body += '2.程式權限申請單申請方式如附件<br>'
  # msg_body += 'P.S.人員名單可依據狀況斟酌增減<br>'
  msg_body += '不好意思剛剛附錯冷軋廠運轉日報表檔案!!!!<br>'
  msg_body += '8224~8226 林育維，系統代號一樣選CRMW<br>PS因冷軋運轉日報表不是EC架構下的程式所以才多寄這封信'
  msg_body += ' <table border=0 bgcolor="#E0E0E0">'
  msg_body += '<br>'
  msg_body += ' <TR bgcolor="#66B3FF">'
  msg_body += '	<TD align="center"  style="width:130"><font color="#FFFFFF"><b>以下是您曾經申請過的產線</font></TD>'
  msg_body += ' </TR>'
  for i in mail[temp]:
    msg_body += ' <TR bgcolor="#FCFCFC">'
    msg_body += '	<TD align="left" >'+i+'</TD>'
    msg_body += ' </TR>'
  msg_body += ' </table>'
  msg_body += '</body></html>'
  recipient = temp+"@mail.yusco.com.tw"
  # recipient = "yu65426@mail.yusco.com.tw"
  subject = "冷軋廠運轉日報表權限申請單"
  rt_code, rt_desc = SEND_MAIL(recipient, subject, msg_body, 'html')
  print('工號:'+temp+'發送成功')
print("郵件發送回傳=>code=" + str(rt_code) + ",  desc=" + rt_desc)
  # print(msg_body)
